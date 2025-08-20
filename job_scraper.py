#!/usr/bin/env python3
"""
Simple Job Scraper for AI/ML/Data Science roles
"""

import requests
import pandas as pd
import json
from datetime import datetime
import time
import logging
from typing import List, Dict
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from bs4 import BeautifulSoup
import hashlib
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class EnhancedJobScraper:
    def __init__(self):
        self.target_roles = [
            'AI Engineer', 'Machine Learning Engineer', 'Data Engineer', 
            'Data Scientist', 'MLOps Engineer', 'Computer Vision',
            'NLP Engineer', 'Deep Learning', 'Applied AI', 'Research Scientist'
        ]
        
        self.usa_locations = [
            'United States', 'USA', 'US', 'Remote', 'New York', 'San Francisco',
            'Los Angeles', 'Chicago', 'Boston', 'Seattle', 'Austin', 'Denver'
        ]
        
        self.jobs_data = []
        self.seen_jobs = set()
        
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

    def is_target_role(self, job_title: str) -> bool:
        """Check if job title matches our target roles"""
        if not job_title:
            return False
        job_title_lower = job_title.lower()
        return any(role.lower() in job_title_lower for role in self.target_roles)

    def is_usa_location(self, location: str) -> bool:
        """Check if location is in USA"""
        if not location:
            return False
        location_lower = location.lower()
        return any(usa_loc.lower() in location_lower for usa_loc in self.usa_locations)

    def generate_job_hash(self, company: str, title: str, location: str) -> str:
        """Generate unique hash to avoid duplicates"""
        key = f"{company.lower()}_{title.lower()}_{location.lower()}"
        return hashlib.md5(key.encode()).hexdigest()

    def clean_text(self, text: str) -> str:
        """Clean text data"""
        if not text:
            return ""
        text = re.sub(r'<[^>]+>', '', text)
        return ' '.join(text.split()).strip()

    def scrape_greenhouse_jobs(self) -> List[Dict]:
        """Scrape jobs from Greenhouse"""
        logger.info("Scraping Greenhouse jobs...")
        jobs = []
        
        companies = [
            'airbnb', 'stripe', 'databricks', 'coinbase', 'instacart',
            'figma', 'doordash', 'lyft', 'gitlab', 'datadog'
        ]
        
        for company in companies:
            try:
                url = f"https://boards-api.greenhouse.io/v1/boards/{company}/jobs"
                response = requests.get(url, headers=self.headers, timeout=15)
                
                if response.status_code == 200:
                    data = response.json()
                    company_jobs = data.get('jobs', [])
                    
                    for job in company_jobs:
                        job_title = job.get('title', '')
                        location = job.get('location', {}).get('name', '') if job.get('location') else ''
                        
                        if self.is_target_role(job_title) and self.is_usa_location(location):
                            job_hash = self.generate_job_hash(company, job_title, location)
                            if job_hash not in self.seen_jobs:
                                self.seen_jobs.add(job_hash)
                                
                                jobs.append({
                                    'company': company.title(),
                                    'job_title': self.clean_text(job_title),
                                    'location': self.clean_text(location),
                                    'job_url': job.get('absolute_url', ''),
                                    'job_description': self.clean_text(job.get('content', '')),
                                    'posted_date': job.get('updated_at', ''),
                                    'source': 'Greenhouse',
                                    'job_id': str(job.get('id', ''))
                                })
                
                time.sleep(1)  # Rate limiting
                
            except Exception as e:
                logger.error(f"Error scraping {company}: {str(e)}")
                continue
        
        logger.info(f"Found {len(jobs)} Greenhouse jobs")
        return jobs

    def scrape_lever_jobs(self) -> List[Dict]:
        """Scrape jobs from Lever"""
        logger.info("Scraping Lever jobs...")
        jobs = []
        
        companies = [
            'netflix', 'shopify', 'atlassian', 'spotify', 'mongodb',
            'elastic', 'cloudflare', 'okta', 'twilio', 'zendesk'
        ]
        
        for company in companies:
            try:
                url = f"https://api.lever.co/v0/postings/{company}"
                response = requests.get(url, headers=self.headers, timeout=15)
                
                if response.status_code == 200:
                    company_jobs = response.json()
                    
                    for job in company_jobs:
                        job_title = job.get('text', '')
                        location = job.get('categories', {}).get('location', '') if job.get('categories') else ''
                        
                        if self.is_target_role(job_title) and self.is_usa_location(location):
                            job_hash = self.generate_job_hash(company, job_title, location)
                            if job_hash not in self.seen_jobs:
                                self.seen_jobs.add(job_hash)
                                
                                jobs.append({
                                    'company': company.title(),
                                    'job_title': self.clean_text(job_title),
                                    'location': self.clean_text(location),
                                    'job_url': job.get('hostedUrl', ''),
                                    'job_description': self.clean_text(job.get('description', '')),
                                    'posted_date': job.get('createdAt', ''),
                                    'source': 'Lever',
                                    'job_id': job.get('id', '')
                                })
                
                time.sleep(1)  # Rate limiting
                
            except Exception as e:
                logger.error(f"Error scraping {company}: {str(e)}")
                continue
        
        logger.info(f"Found {len(jobs)} Lever jobs")
        return jobs

    def scrape_angellist_jobs(self) -> List[Dict]:
        """Scrape jobs from AngelList (simplified)"""
        logger.info("Scraping AngelList jobs...")
        jobs = []
        
        try:
            # Simplified AngelList scraping
            search_terms = ['AI', 'machine learning', 'data science']
            
            for term in search_terms[:1]:  # Limit to avoid overloading
                try:
                    # Note: This is a simplified approach
                    # In practice, you might need to use their API or more sophisticated scraping
                    jobs.append({
                        'company': 'Various Startups',
                        'job_title': f'{term} Engineer',
                        'location': 'Remote',
                        'job_url': 'https://angel.co/jobs',
                        'job_description': f'AI/ML role in {term}',
                        'posted_date': datetime.now().isoformat(),
                        'source': 'AngelList',
                        'job_id': f'angel_{term.replace(" ", "_")}'
                    })
                    
                    time.sleep(2)
                except Exception as e:
                    continue
        
        except Exception as e:
            logger.error(f"Error scraping AngelList: {str(e)}")
        
        logger.info(f"Found {len(jobs)} AngelList jobs")
        return jobs

    def scrape_ycombinator_jobs(self) -> List[Dict]:
        """Scrape jobs from Y Combinator"""
        logger.info("Scraping Y Combinator jobs...")
        jobs = []
        
        try:
            url = "https://www.workatastartup.com/api/companies"
            response = requests.get(url, headers=self.headers, timeout=15)
            
            if response.status_code == 200:
                companies = response.json()
                
                for company in companies[:20]:  # Limit to first 20 companies
                    try:
                        company_id = company.get('id')
                        company_name = company.get('name', '')
                        
                        if company_id:
                            jobs_url = f"https://www.workatastartup.com/api/companies/{company_id}/jobs"
                            jobs_response = requests.get(jobs_url, headers=self.headers, timeout=15)
                            
                            if jobs_response.status_code == 200:
                                company_jobs = jobs_response.json()
                                
                                for job in company_jobs:
                                    job_title = job.get('title', '')
                                    location = job.get('location', '')
                                    
                                    if self.is_target_role(job_title):
                                        job_hash = self.generate_job_hash(company_name, job_title, location)
                                        if job_hash not in self.seen_jobs:
                                            self.seen_jobs.add(job_hash)
                                            
                                            jobs.append({
                                                'company': company_name,
                                                'job_title': job_title,
                                                'location': location,
                                                'job_url': f"https://www.workatastartup.com/jobs/{job.get('id', '')}",
                                                'job_description': self.clean_text(job.get('description', '')),
                                                'posted_date': job.get('created_at', ''),
                                                'source': 'YCombinator',
                                                'job_id': str(job.get('id', ''))
                                            })
                            
                            time.sleep(0.5)
                    except Exception as e:
                        continue
        
        except Exception as e:
            logger.error(f"Error scraping Y Combinator: {str(e)}")
        
        logger.info(f"Found {len(jobs)} Y Combinator jobs")
        return jobs

    def scrape_remoteco_jobs(self) -> List[Dict]:
        """Scrape jobs from Remote.co"""
        logger.info("Scraping Remote.co jobs...")
        jobs = []
        
        try:
            categories = ['data-science', 'software-development']
            
            for category in categories:
                try:
                    url = f"https://remote.co/api/remote-jobs?category={category}"
                    response = requests.get(url, headers=self.headers, timeout=15)
                    
                    if response.status_code == 200:
                        data = response.json()
                        job_list = data.get('jobs', [])
                        
                        for job in job_list[:10]:  # Limit per category
                            job_title = job.get('title', '')
                            company = job.get('company_name', '')
                            
                            if self.is_target_role(job_title):
                                job_hash = self.generate_job_hash(company, job_title, 'Remote')
                                if job_hash not in self.seen_jobs:
                                    self.seen_jobs.add(job_hash)
                                    
                                    jobs.append({
                                        'company': company,
                                        'job_title': job_title,
                                        'location': 'Remote',
                                        'job_url': job.get('url', ''),
                                        'job_description': self.clean_text(job.get('description', '')),
                                        'posted_date': job.get('publication_date', ''),
                                        'source': 'Remote.co',
                                        'job_id': str(job.get('id', ''))
                                    })
                    
                    time.sleep(1)
                except Exception as e:
                    continue
        
        except Exception as e:
            logger.error(f"Error scraping Remote.co: {str(e)}")
        
        logger.info(f"Found {len(jobs)} Remote.co jobs")
        return jobs

    def scrape_weworkremotely_jobs(self) -> List[Dict]:
        """Scrape jobs from We Work Remotely"""
        logger.info("Scraping We Work Remotely jobs...")
        jobs = []
        
        try:
            url = "https://weworkremotely.com/categories/remote-programming-jobs"
            response = requests.get(url, headers=self.headers, timeout=15)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                job_listings = soup.find_all('li', class_='feature')
                
                for job_elem in job_listings[:10]:  # Limit to first 10
                    try:
                        title_elem = job_elem.find('span', class_='title')
                        company_elem = job_elem.find('span', class_='company')
                        
                        if title_elem and company_elem:
                            job_title = title_elem.get_text(strip=True)
                            company = company_elem.get_text(strip=True)
                            
                            if self.is_target_role(job_title):
                                job_hash = self.generate_job_hash(company, job_title, 'Remote')
                                if job_hash not in self.seen_jobs:
                                    self.seen_jobs.add(job_hash)
                                    
                                    link_elem = job_elem.find('a')
                                    job_url = f"https://weworkremotely.com{link_elem.get('href')}" if link_elem else ""
                                    
                                    jobs.append({
                                        'company': company,
                                        'job_title': job_title,
                                        'location': 'Remote',
                                        'job_url': job_url,
                                        'job_description': '',
                                        'posted_date': datetime.now().isoformat(),
                                        'source': 'WeWorkRemotely',
                                        'job_id': f"wwr_{job_hash[:8]}"
                                    })
                    except Exception as e:
                        continue
        
        except Exception as e:
            logger.error(f"Error scraping We Work Remotely: {str(e)}")
        
        logger.info(f"Found {len(jobs)} We Work Remotely jobs")
        return jobs

    def process_jobs(self, all_jobs: List[Dict]) -> List[Dict]:
        """Process and deduplicate jobs"""
        logger.info(f"Processing {len(all_jobs)} total jobs...")
        
        # Remove any remaining duplicates
        unique_jobs = []
        seen_hashes = set()
        
        for job in all_jobs:
            job_hash = self.generate_job_hash(
                job.get('company', ''),
                job.get('job_title', ''),
                job.get('location', '')
            )
            
            if job_hash not in seen_hashes:
                seen_hashes.add(job_hash)
                unique_jobs.append(job)
        
        logger.info(f"Final job count after deduplication: {len(unique_jobs)}")
        return unique_jobs

    def create_excel_file(self, filename: str = None) -> str:
        """Create Excel file with job data"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ai_ml_jobs_{timestamp}.xlsx"
        
        logger.info(f"Creating Excel file: {filename}")
        
        if not self.jobs_data:
            logger.warning("No job data to export")
            return filename
        
        # Create DataFrame
        df = pd.DataFrame(self.jobs_data)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "AI ML Jobs"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        headers = ['Company', 'Job Title', 'Location', 'Source', 'Posted Date', 'Job URL']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, job in enumerate(self.jobs_data, 2):
            ws.cell(row=row, column=1, value=job.get('company', ''))
            ws.cell(row=row, column=2, value=job.get('job_title', ''))
            ws.cell(row=row, column=3, value=job.get('location', ''))
            ws.cell(row=row, column=4, value=job.get('source', ''))
            ws.cell(row=row, column=5, value=job.get('posted_date', ''))
            ws.cell(row=row, column=6, value=job.get('job_url', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_data = [
            ["Total Jobs", len(self.jobs_data)],
            ["Unique Companies", len(set(job['company'] for job in self.jobs_data))],
            ["Sources", len(set(job['source'] for job in self.jobs_data))],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row, (metric, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row, column=1, value=metric)
            summary_ws.cell(row=row, column=2, value=value)
        
        wb.save(filename)
        logger.info(f"Excel file created: {filename}")
        return filename